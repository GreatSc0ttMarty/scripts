from openpyxl import load_workbook, Workbook
import openpyxl

print('\nLoading files..')

file1 = '{enter file path here}'
file2 = '{enter second file path here}'

wb = load_workbook(filename=f"{file1}.xlsx")
wb2 = Workbook()
wb2_2 = wb2.active
wb2.save(filename=f"{file2}.xlsx")
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']


print('\Copying contents to variables..')

def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]


sheet1list = [x for x in iter_rows(sheet1)]
sheet2list = [x for x in iter_rows(sheet2)]


print('\nPlease wait.. Merging data..')

# Make your own cell sorting algorithim

for count, row in enumerate(sheet1list):
    for row2 in sheet2list:
        if row[1] == str(row2[1]) + ' ' + str(row2[2]):
            #print(row[1], '=', str(row2[1]) + ' ' + str(row2[2]))
            row2[5] = row[2]
            row2[6] = row[4]
            row2[7] = row[3]
            row2[8] = row[5]
            row2[9] = row[3]


print('\nApplying to file!')


for count, row in enumerate(sheet2list):
    for i in range(0,10):
        wb2_2.cell(row=count+1, column=i+1).value = sheet2list[count][i]

print('\nSaving file..')

wb2.save(f'{file2}.xlsx')

print('\Complete!')